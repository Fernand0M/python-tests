import os
from openpyxl import load_workbook


class ContractsBook():
    """ Load and store data from the contract list excel book """

    def __init__(self, file_path):
        self.__sheet_name = 'Data'
        self.__name_index = 'A'
        self.__dob_index = 'B'
        self.__addr_index = 'C'
        self.__cwd = os.getcwd()

        self.wb = None
        self.sheet = None
        self.file_path = file_path
        self.data = []

        self.load_workbook()
        self.read_data()

    def load_workbook(self):
        self.wb = load_workbook(self.__cwd + self.file_path)
        if self.wb is None:
            raise IOError
        else:
            self.sheet = self.wb['Data']

    def read_data(self):
        i = 2  # data starts at row 2
        while self.has_data(i):
            contract = {}
            contract['name'] = self.get_cell_value(self.__name_index, i)
            contract['dob'] = self.get_cell_value(self.__dob_index, i)
            contract['address'] = self.get_cell_value(self.__addr_index, i)
            self.data.append(contract)
            i += 1

    def has_data(self, index):
        return self.get_cell_value(self.__name_index, index)

    def get_cell_value(self, char, index):
        return self.sheet[char + str(index)].value


if __name__ == '__main__':
    cb = ContractsBook('/data/data.xlsx')
